import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment
import numpy as np

# 创建示例数据
data = {
    'Name': ['Alice', 'Alice', 'Bob', np.nan , 'Charlie','LA','LB','LA'],
    'Age': [25, 25, 30, 30, 35, 30, 30, 30],
    'Location': ['NY', 'NY', 'LA', 'LA', 'Chicago','LA',np.nan,'LA']
}

df = pd.DataFrame(data)

# 导出到Excel文件
excel_path = 'test.xlsx'
df.to_excel(excel_path, index=False)

# 加载Excel文件
wb = load_workbook(excel_path)
ws = wb.active

# 定义一个函数来合并相同的单元格
def merge_cells(ws, start_row, end_row, col):
    cell_value = ws.cell(start_row, col).value
    if all(ws.cell(row, col).value == cell_value for row in range(start_row, end_row + 1)):
        ws.merge_cells(start_row=start_row, start_column=col, end_row=end_row, end_column=col)
        ws.cell(start_row, col).alignment = Alignment(vertical='center')

# 获取数据的范围
max_row = ws.max_row
max_col = ws.max_column

# 遍历每一列，找到相同单元格并合并
for col in range(1, max_col + 1):
    start_row = 2
    for row in range(2, max_row + 1):
        if ws.cell(row, col).value != ws.cell(row + 1, col).value:
            merge_cells(ws, start_row, row, col)
            start_row = row + 1

# 保存Excel文件
wb.save(excel_path)

print(f"Excel文件已成功保存到 {excel_path}")
